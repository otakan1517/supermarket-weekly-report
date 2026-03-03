from __future__ import annotations

import argparse
from pathlib import Path
from datetime import datetime as dt

import pandas as pd
import yaml
from openpyxl import load_workbook

SHEET_README = "00_readme"
SHEET_SUMMARY = "01_summary"
SHEET_STORE = "02_store"
SHEET_DEPT_CAT = "03_dept_category"
SHEET_TREND = "04_trend_daily"
SHEET_QUALITY = "90_quality"

REQUIRED = ["datetime", "store", "department", "category", "qty", "sales_yen", "cost_yen"]


def load_config(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def apply_column_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    # mapping: canonical -> actual
    rename = {}
    for canonical, actual in mapping.items():
        if actual in df.columns and canonical != actual:
            rename[actual] = canonical
    if rename:
        df = df.rename(columns=rename)
    return df


def add_time_fields(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
    df["date"] = df["datetime"].dt.date

    iso = df["datetime"].dt.isocalendar()
    df["iso_year"] = iso["year"].astype("Int64")
    df["iso_week"] = iso["week"].astype("Int64")
    df["week_label"] = df["iso_year"].astype(str) + "-W" + df["iso_week"].astype(int).astype(str).str.zfill(2)
    return df


def validate(df: pd.DataFrame) -> pd.DataFrame:
    issues = []

    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        issues.append({"severity": "ERROR", "check": "missing_required_columns", "count": len(missing),
                       "note": f"必須列が不足: {missing}", "action": "入力列名 or config.yml のcolumnsを確認"})
        return pd.DataFrame(issues)

    # datetime パース
    dt_na = int(df["datetime"].isna().sum())
    if dt_na > 0:
        sev = "ERROR" if dt_na == len(df) else "WARN"
        issues.append({"severity": sev, "check": "datetime_parse_failed", "count": dt_na,
                       "note": "datetimeがNaN（パース失敗）", "action": "datetime形式を確認"})

    # 欠損（必須）
    for c in REQUIRED:
        n = int(df[c].isna().sum())
        if n > 0:
            issues.append({"severity": "WARN", "check": f"missing_{c}", "count": n,
                           "note": f"{c} に欠損", "action": "欠損補完/除外の方針を決める"})

    # 負値
    for c in ["qty", "sales_yen", "cost_yen", "discount_yen"]:
        if c in df.columns:
            n = int((df[c] < 0).sum())
            if n > 0:
                issues.append({"severity": "WARN", "check": f"negative_{c}", "count": n,
                               "note": f"{c} が負値", "action": "返品/取消の扱いを決める"})

    # 原価>売上
    n = int((df["cost_yen"] > df["sales_yen"]).sum())
    if n > 0:
        issues.append({"severity": "WARN", "check": "cost_gt_sales", "count": n,
                       "note": "原価が売上を上回る行", "action": "原価/売価の定義や値引きを確認"})

    # 重複参考
    if "transaction_id" in df.columns:
        key = ["transaction_id", "datetime"]
        if "product_code" in df.columns:
            key.append("product_code")
        dup = int(df.duplicated(subset=key, keep=False).sum())
        if dup > 0:
            issues.append({"severity": "INFO", "check": "possible_duplicates", "count": dup,
                           "note": f"重複っぽい（キー={key}）", "action": "必要なら重複定義を見直す"})

    if not issues:
        issues = [{"severity": "INFO", "check": "ok", "count": 0, "note": "問題なし", "action": ""}]

    return pd.DataFrame(issues)


def agg_kpi(g: pd.DataFrame) -> pd.Series:
    sales = float(g["sales_yen"].sum())
    cost = float(g["cost_yen"].sum())
    gp = float((g["sales_yen"] - g["cost_yen"]).sum())
    qty = float(g["qty"].sum())
    disc = float(g["discount_yen"].sum()) if "discount_yen" in g.columns else 0.0
    margin = gp / sales if sales != 0 else pd.NA
    return pd.Series({
        "sales_yen": sales,
        "cost_yen": cost,
        "gross_profit": gp,
        "gross_margin": margin,
        "qty": qty,
        "discount_yen": disc,
    })


def pick_target_week(df: pd.DataFrame, target: str) -> str:
    weeks = sorted([w for w in df["week_label"].dropna().unique().tolist() if w != "NaT-Wnan"])
    if target and target in weeks:
        return target
    return weeks[-1] if weeks else ""


def post_format(xlsx_path: Path, sheet_names: list[str]) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    # 表示形式
    FMT_YEN = '¥#,##0'
    FMT_PCT = '0.00%'
    FMT_INT = '#,##0'

    MONEY_COLS = {"sales_yen", "cost_yen", "gross_profit", "discount_yen"}
    PCT_COLS = {"gross_margin"}
    INT_COLS = {"qty", "sales_rank", "count"}

    # 列幅（ヘッダー名ベース）
    WIDTH_BY_SHEET = {
        "01_summary": {
            "week_label": 12, "sales_yen": 14, "cost_yen": 14, "gross_profit": 14,
            "gross_margin": 10, "qty": 10, "discount_yen": 14
        },
        "02_store": {
            "week_label": 12, "store": 14, "sales_yen": 14, "cost_yen": 14, "gross_profit": 14,
            "gross_margin": 10, "qty": 10, "discount_yen": 14, "sales_rank": 10
        },
        "03_dept_category": {
            "week_label": 12, "department": 12, "category": 12, "sales_yen": 14, "cost_yen": 14,
            "gross_profit": 14, "gross_margin": 10, "qty": 10, "discount_yen": 14
        },
        "04_trend_daily": {
            "week_label": 12, "date": 12, "sales_yen": 14, "cost_yen": 14, "gross_profit": 14,
            "gross_margin": 10, "qty": 10, "discount_yen": 14, "is_holiday": 10, "holiday_name": 18
        },
        "90_quality": {
            "severity": 10, "check": 24, "count": 10, "note": 30, "action": 40
        },
    }

    wb = load_workbook(xlsx_path)

    def apply_formats(ws):
        headers = [c.value for c in ws[1]]
        idx = {h: i + 1 for i, h in enumerate(headers) if isinstance(h, str)}

        for h in (MONEY_COLS & idx.keys()):
            c = idx[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = FMT_YEN

        for h in (PCT_COLS & idx.keys()):
            c = idx[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = FMT_PCT

        for h in (INT_COLS & idx.keys()):
            c = idx[h]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = FMT_INT

        return idx

    def set_widths(ws, sheet_name, idx):
        mapping = WIDTH_BY_SHEET.get(sheet_name, {})
        for col_name, w in mapping.items():
            if col_name in idx:
                letter = get_column_letter(idx[col_name])
                ws.column_dimensions[letter].width = w  # 列幅設定（openpyxl） citeturn0search3

    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]

        # ヘッダー固定
        ws.freeze_panes = "A2"

        # フィルタ
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        # 00_readme：読みやすく（列幅＋折り返し）
        if name == "00_readme":
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 70
            for cell in ws["B"]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            continue

        # 数値整形 + 列幅
        idx = apply_formats(ws)
        set_widths(ws, name, idx)

    wb.save(xlsx_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="configs/config.yml")
    ap.add_argument("--in", dest="inp", default=None)
    ap.add_argument("--out", dest="out", default=None)
    args = ap.parse_args()

    cfg = load_config(args.config)
    input_path = args.inp or cfg.get("input_path")
    output_dir = Path(cfg.get("output_dir", "outputs"))
    output_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_excel(input_path)
    df = apply_column_mapping(df, cfg.get("columns", {}))
    df = add_time_fields(df)

    qdf = validate(df)
    fatal = bool((qdf["severity"] == "ERROR").any())
    if fatal:
        # ERRORなら出力は作らず止める（仕様どおり）
        print("ERROR: 品質チェックで致命的な問題が見つかりました。90_quality 相当を確認してください。")
        print(qdf)
        raise SystemExit(1)

    target_week = pick_target_week(df, cfg.get("target_week_label", ""))
    if not target_week:
        raise ValueError("week_label を特定できません（datetimeの欠損が多い可能性）")

    dfw = df[df["week_label"] == target_week].copy()

    # 集計
    summary = dfw.groupby(["week_label"]).apply(agg_kpi).reset_index()

    store = dfw.groupby(["week_label", "store"]).apply(agg_kpi).reset_index()
    store = store.sort_values(["week_label", "sales_yen"], ascending=[True, False])
    store["sales_rank"] = store.groupby("week_label")["sales_yen"].rank(ascending=False, method="dense").astype(int)

    dept_cat = dfw.groupby(["week_label", "department", "category"]).apply(agg_kpi).reset_index()
    dept_cat = dept_cat.sort_values(["week_label", "sales_yen"], ascending=[True, False])

    trend = dfw.groupby(["week_label", "date"]).apply(agg_kpi).reset_index().sort_values(["week_label", "date"])
    if "is_holiday" in dfw.columns:
        hol = dfw.groupby(["week_label", "date"], as_index=False)["is_holiday"].max()
        trend = trend.merge(hol, on=["week_label", "date"], how="left")
    if "holiday_name" in dfw.columns:
        hn = dfw[dfw["holiday_name"].fillna("") != ""].groupby(["week_label", "date"], as_index=False)["holiday_name"].first()
        trend = trend.merge(hn, on=["week_label", "date"], how="left")

    readme = pd.DataFrame([
        {"key": "generated_at", "value": dt.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"key": "input_path", "value": str(input_path)},
        {"key": "week_label", "value": target_week},
        {"key": "rows", "value": len(dfw)},
        {"key": "notes", "value": "90_quality を確認し、WARNがあれば入力定義・欠損・異常値を見直してください"},
    ])

    out_path = Path(args.out) if args.out else (output_dir / f"weekly_report_{target_week}.xlsx")

    with pd.ExcelWriter(out_path, engine="openpyxl") as w:
        readme.to_excel(w, sheet_name=SHEET_README, index=False)
        summary.to_excel(w, sheet_name=SHEET_SUMMARY, index=False)
        store.to_excel(w, sheet_name=SHEET_STORE, index=False)
        dept_cat.to_excel(w, sheet_name=SHEET_DEPT_CAT, index=False)
        trend.to_excel(w, sheet_name=SHEET_TREND, index=False)
        qdf.to_excel(w, sheet_name=SHEET_QUALITY, index=False)

    post_format(out_path, [SHEET_README, SHEET_SUMMARY, SHEET_STORE, SHEET_DEPT_CAT, SHEET_TREND, SHEET_QUALITY])
    print(f"saved: {out_path}")


if __name__ == "__main__":
    main()
